from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Request, Response
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
import json
import tempfile
import uuid
import difflib
import httpx
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from motor.motor_asyncio import AsyncIOMotorClient

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Temp storage for uploaded files and generated reports
TEMP_DIR = Path(tempfile.gettempdir()) / "json_compare"
TEMP_DIR.mkdir(exist_ok=True)

# Max file size: 30 MB
MAX_FILE_SIZE = 30 * 1024 * 1024

# MongoDB setup
MONGO_URL = os.environ.get('MONGO_URL')
DB_NAME = os.environ.get('DB_NAME', 'test_database')
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============== MODELS ==============

class JsonUploadResponse(BaseModel):
    file_id: str
    filename: str
    size: int
    valid: bool
    error: Optional[str] = None
    structure: Optional[Dict[str, Any]] = None

class ToolPath(BaseModel):
    path: List[str]
    path_string: str
    tool_count: int

class AnalyzeResponse(BaseModel):
    detected_paths: List[ToolPath]
    json_structure: Dict[str, Any]

class Tool(BaseModel):
    name: str
    description: str
    index: int

class CompareRequest(BaseModel):
    file1_id: str
    file2_id: str
    compare_type: str
    custom_path: Optional[str] = None
    selected_tools: Optional[List[str]] = None

class ComparisonSummary(BaseModel):
    file1_tools: int
    file2_tools: int
    same_count: int
    modified_count: int
    added_count: int
    removed_count: int
    excel_filename: str
    download_url: str
    preview_data: Optional[Dict[str, Any]] = None

class HistoryItem(BaseModel):
    id: str
    file1_name: str
    file2_name: str
    compare_type: str
    timestamp: str
    file1_tools: int
    file2_tools: int
    same_count: int
    modified_count: int
    added_count: int
    removed_count: int
    preview_data: Optional[Dict[str, Any]] = None

class SaveHistoryRequest(BaseModel):
    file1_name: str
    file2_name: str
    compare_type: str
    file1_tools: int
    file2_tools: int
    same_count: int
    modified_count: int
    added_count: int
    removed_count: int
    preview_data: Dict[str, Any]

class GoogleSheetsExportRequest(BaseModel):
    access_token: str
    spreadsheet_title: str
    preview_data: Dict[str, Any]

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None

# ============== AUTH FUNCTIONS ==============

async def get_current_user(request: Request) -> Optional[User]:
    """Get current user from session token in cookie or Authorization header."""
    session_token = request.cookies.get("session_token")
    
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header[7:]
    
    if not session_token:
        return None
    
    try:
        session_doc = await db.user_sessions.find_one(
            {"session_token": session_token},
            {"_id": 0}
        )
        
        if not session_doc:
            return None
        
        expires_at = session_doc.get("expires_at")
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if expires_at < datetime.now(timezone.utc):
            return None
        
        user_doc = await db.users.find_one(
            {"user_id": session_doc["user_id"]},
            {"_id": 0}
        )
        
        if not user_doc:
            return None
        
        return User(**user_doc)
    except Exception as e:
        logger.error(f"Auth error: {e}")
        return None

# ============== UTILITY FUNCTIONS ==============

def get_json_structure(data: Any, max_depth: int = 10, current_depth: int = 0) -> Dict[str, Any]:
    """Generate a simplified structure representation of JSON data."""
    if current_depth >= max_depth:
        return {"type": "...", "truncated": True}
    
    if isinstance(data, dict):
        return {
            "type": "object",
            "keys": {k: get_json_structure(v, max_depth, current_depth + 1) for k, v in list(data.items())[:50]}
        }
    elif isinstance(data, list):
        if len(data) > 0:
            return {
                "type": "array",
                "length": len(data),
                "sample": get_json_structure(data[0], max_depth, current_depth + 1) if len(data) > 0 else None
            }
        return {"type": "array", "length": 0}
    elif isinstance(data, str):
        return {"type": "string", "preview": data[:100] if len(data) > 100 else data}
    elif isinstance(data, bool):
        return {"type": "boolean", "value": data}
    elif isinstance(data, (int, float)):
        return {"type": "number", "value": data}
    elif data is None:
        return {"type": "null"}
    else:
        return {"type": str(type(data).__name__)}

def find_array_paths(data: Any, current_path: List[str] = None, results: List[ToolPath] = None) -> List[ToolPath]:
    """Find all paths to arrays of objects in the JSON."""
    if current_path is None:
        current_path = []
    if results is None:
        results = []
    
    if isinstance(data, dict):
        for key, value in data.items():
            find_array_paths(value, current_path + [key], results)
    elif isinstance(data, list) and len(data) > 0:
        if isinstance(data[0], dict):
            results.append(ToolPath(
                path=current_path,
                path_string=" -> ".join(current_path) if current_path else "root",
                tool_count=len(data)
            ))
        for item in data[:1]:
            if isinstance(item, dict):
                for key, value in item.items():
                    find_array_paths(value, current_path + ["[0]", key], results)
    
    return results

TOOL_PATHS = [
    ["log", "body", "toolConfig", "tools"],
    ["log", "body", "tools"],
    ["body", "toolConfig", "tools"],
    ["body", "tools"],
    ["toolConfig", "tools"],
    ["tools"]
]

def extract_tools(data: Dict, custom_path: Optional[str] = None) -> Tuple[List[Dict], str]:
    """Extract tools from JSON using predefined or custom path."""
    paths_to_try = []
    
    if custom_path:
        if " -> " in custom_path:
            paths_to_try = [custom_path.split(" -> ")]
        else:
            paths_to_try = [custom_path.split(".")]
    else:
        paths_to_try = TOOL_PATHS
    
    for path in paths_to_try:
        try:
            current = data
            for key in path:
                current = current[key]
            
            if isinstance(current, list) and len(current) > 0:
                return current, " -> ".join(path)
        except (KeyError, TypeError, IndexError):
            continue
    
    return [], ""

def normalize_tool(tool: Dict, index: int) -> Dict:
    """Normalize tool data to handle different field names."""
    name = (
        tool.get("name") or 
        tool.get("toolSpec", {}).get("name") or
        tool.get("tool_name") or
        tool.get("function", {}).get("name") or
        f"Tool_{index}"
    )
    
    description = (
        tool.get("description") or
        tool.get("toolSpec", {}).get("description") or
        tool.get("tool_description") or
        tool.get("function", {}).get("description") or
        json.dumps(tool, indent=2)
    )
    
    return {
        "name": name,
        "description": description,
        "raw": tool
    }

def get_word_diff(text1: str, text2: str) -> Tuple[List[Dict], List[Dict]]:
    """
    Get word-level diff between two texts.
    Returns lists of {text, type} where type is 'same', 'added', or 'removed'.
    Compares word-by-word including spaces, special characters, and numbers.
    """
    # Check if texts are identical
    if text1 == text2:
        return ([{"text": text1, "type": "same"}] if text1 else [], 
                [{"text": text2, "type": "same"}] if text2 else [])
    
    # Split into words (preserves special characters and numbers in words)
    words1 = text1.split() if text1 else []
    words2 = text2.split() if text2 else []
    
    # Use SequenceMatcher for word-by-word comparison
    matcher = difflib.SequenceMatcher(None, words1, words2)
    
    diff1 = []  # For file1 (shows removals)
    diff2 = []  # For file2 (shows additions)
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            # Same words in both
            text = ' '.join(words1[i1:i2])
            if text:
                diff1.append({"text": text, "type": "same"})
                diff2.append({"text": text, "type": "same"})
        elif tag == 'delete':
            # Words only in file1 (removed)
            text = ' '.join(words1[i1:i2])
            if text:
                diff1.append({"text": text, "type": "removed"})
        elif tag == 'insert':
            # Words only in file2 (added)
            text = ' '.join(words2[j1:j2])
            if text:
                diff2.append({"text": text, "type": "added"})
        elif tag == 'replace':
            # Different words
            text1_part = ' '.join(words1[i1:i2])
            text2_part = ' '.join(words2[j1:j2])
            if text1_part:
                diff1.append({"text": text1_part, "type": "removed"})
            if text2_part:
                diff2.append({"text": text2_part, "type": "added"})
    
    return diff1, diff2

def get_text_diff_detailed(text1: str, text2: str) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str]]]:
    """Compare two texts and return word-level chunks with their status."""
    words1 = text1.split()
    words2 = text2.split()
    
    matcher = difflib.SequenceMatcher(None, words1, words2)
    
    chunks1 = []
    chunks2 = []
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            text = ' '.join(words1[i1:i2])
            if text:
                chunks1.append((text + ' ', 'same'))
                chunks2.append((text + ' ', 'same'))
        elif tag == 'delete':
            text = ' '.join(words1[i1:i2])
            if text:
                chunks1.append((text + ' ', 'removed'))
        elif tag == 'insert':
            text = ' '.join(words2[j1:j2])
            if text:
                chunks2.append((text + ' ', 'added'))
        elif tag == 'replace':
            text1_part = ' '.join(words1[i1:i2])
            text2_part = ' '.join(words2[j1:j2])
            if text1_part:
                chunks1.append((text1_part + ' ', 'removed'))
            if text2_part:
                chunks2.append((text2_part + ' ', 'added'))
    
    return chunks1, chunks2

def create_excel_comparison(tools1: List[Dict], tools2: List[Dict], 
                           selected_tools: Optional[List[str]], output_path: str) -> Dict:
    """Create Excel file with comparison of tools."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Normalize tools
    normalized_tools1 = [normalize_tool(t, i) for i, t in enumerate(tools1, 1)]
    normalized_tools2 = [normalize_tool(t, i) for i, t in enumerate(tools2, 1)]
    
    # Filter by selected tools if specified
    if selected_tools:
        normalized_tools1 = [t for t in normalized_tools1 if t["name"] in selected_tools]
        normalized_tools2 = [t for t in normalized_tools2 if t["name"] in selected_tools]
    
    # Create sheets
    ws_comparison = wb.create_sheet("Comparison")
    ws_differences = wb.create_sheet("Differences")
    ws_file1 = wb.create_sheet("File1_Tools")
    ws_file2 = wb.create_sheet("File2_Tools")
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    removal_fill = PatternFill(start_color="FFD7D5", end_color="FFD7D5", fill_type="solid")
    addition_fill = PatternFill(start_color="D4F4DD", end_color="D4F4DD", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Build tool dictionaries
    tools1_dict = {t["name"]: t["description"] for t in normalized_tools1}
    tools2_dict = {t["name"]: t["description"] for t in normalized_tools2}
    
    all_names = sorted(set(tools1_dict.keys()) | set(tools2_dict.keys()))
    
    # Stats
    same_count = 0
    modified_count = 0
    added_count = 0
    removed_count = 0
    
    # === COMPARISON SHEET ===
    headers = ["Tool Name", "In File1", "In File2", "Description Same?", "Notes"]
    ws_comparison.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_comparison.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    row_num = 2
    for name in all_names:
        in_file1 = "✓" if name in tools1_dict else "✗"
        in_file2 = "✓" if name in tools2_dict else "✗"
        
        if name in tools1_dict and name in tools2_dict:
            desc1 = tools1_dict[name].strip()
            desc2 = tools2_dict[name].strip()
            desc_same = "✓" if desc1 == desc2 else "✗"
            
            if desc1 == desc2:
                notes = "Same in both"
                same_count += 1
            else:
                notes = "Description differs"
                modified_count += 1
        else:
            desc_same = "N/A"
            if name not in tools1_dict:
                notes = "Only in File2"
                added_count += 1
            else:
                notes = "Only in File1"
                removed_count += 1
        
        ws_comparison.append([name, in_file1, in_file2, desc_same, notes])
        
        for col_num in range(1, 6):
            cell = ws_comparison.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            if col_num == 2:
                cell.fill = green_fill if in_file1 == "✓" else red_fill
            elif col_num == 3:
                cell.fill = green_fill if in_file2 == "✓" else red_fill
            elif col_num == 4:
                if desc_same == "✓":
                    cell.fill = green_fill
                elif desc_same == "✗":
                    cell.fill = yellow_fill
        
        row_num += 1
    
    ws_comparison.column_dimensions['A'].width = 35
    ws_comparison.column_dimensions['B'].width = 20
    ws_comparison.column_dimensions['C'].width = 20
    ws_comparison.column_dimensions['D'].width = 20
    ws_comparison.column_dimensions['E'].width = 30
    
    # === DIFFERENCES SHEET ===
    diff_headers = ["Tool Name", "File1 Description (Removals)", "File2 Description (Additions)", "Change Type"]
    ws_differences.append(diff_headers)
    
    for col_num, header in enumerate(diff_headers, 1):
        cell = ws_differences.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    diff_row_num = 2
    
    for name in all_names:
        desc1 = tools1_dict.get(name, "")
        desc2 = tools2_dict.get(name, "")
        
        if desc1.strip() != desc2.strip():
            if not desc1:
                change_type = "Added in File2"
            elif not desc2:
                change_type = "Removed from File2"
            else:
                change_type = "Modified"
            
            if desc1 and desc2:
                chunks1, chunks2 = get_text_diff_detailed(desc1, desc2)
            else:
                chunks1 = [(desc1, 'removed')] if desc1 else []
                chunks2 = [(desc2, 'added')] if desc2 else []
            
            ws_differences.append([name, "", "", change_type])
            
            name_cell = ws_differences.cell(row=diff_row_num, column=1)
            name_cell.value = name
            name_cell.border = border
            name_cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            cell_file1 = ws_differences.cell(row=diff_row_num, column=2)
            if chunks1:
                has_removals = any(status == 'removed' for _, status in chunks1)
                cell_file1.value = ''.join(text for text, _ in chunks1).strip()
                if has_removals:
                    cell_file1.fill = removal_fill
            cell_file1.border = border
            cell_file1.alignment = Alignment(vertical="top", wrap_text=True)
            
            cell_file2 = ws_differences.cell(row=diff_row_num, column=3)
            if chunks2:
                has_additions = any(status == 'added' for _, status in chunks2)
                cell_file2.value = ''.join(text for text, _ in chunks2).strip()
                if has_additions:
                    cell_file2.fill = addition_fill
            cell_file2.border = border
            cell_file2.alignment = Alignment(vertical="top", wrap_text=True)
            
            change_cell = ws_differences.cell(row=diff_row_num, column=4)
            change_cell.value = change_type
            change_cell.border = border
            change_cell.alignment = Alignment(horizontal="center", vertical="center")
            if change_type == "Added in File2":
                change_cell.fill = addition_fill
            elif change_type == "Removed from File2":
                change_cell.fill = removal_fill
            else:
                change_cell.fill = yellow_fill
            
            diff_row_num += 1
    
    ws_differences.column_dimensions['A'].width = 35
    ws_differences.column_dimensions['B'].width = 70
    ws_differences.column_dimensions['C'].width = 70
    ws_differences.column_dimensions['D'].width = 20
    
    for row in range(2, diff_row_num):
        ws_differences.row_dimensions[row].height = 100
    
    # === FILE 1 DETAILS SHEET ===
    ws_file1.append(["#", "Tool Name", "Description"])
    for i in range(1, 4):
        cell = ws_file1.cell(row=1, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, tool in enumerate(normalized_tools1, 1):
        ws_file1.append([idx, tool["name"], tool["description"]])
    
    for row in ws_file1.iter_rows(min_row=1, max_row=ws_file1.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    ws_file1.column_dimensions['A'].width = 5
    ws_file1.column_dimensions['B'].width = 35
    ws_file1.column_dimensions['C'].width = 100
    
    # === FILE 2 DETAILS SHEET ===
    ws_file2.append(["#", "Tool Name", "Description"])
    for i in range(1, 4):
        cell = ws_file2.cell(row=1, column=i)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, tool in enumerate(normalized_tools2, 1):
        ws_file2.append([idx, tool["name"], tool["description"]])
    
    for row in ws_file2.iter_rows(min_row=1, max_row=ws_file2.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    ws_file2.column_dimensions['A'].width = 5
    ws_file2.column_dimensions['B'].width = 35
    ws_file2.column_dimensions['C'].width = 100
    
    wb.save(output_path)
    
    # Build preview data with word-level diffs
    comparison_data = []
    differences_data = []
    file1_tools_data = []
    file2_tools_data = []
    
    for name in all_names:
        in_file1 = name in tools1_dict
        in_file2 = name in tools2_dict
        
        if in_file1 and in_file2:
            desc1 = tools1_dict[name].strip()
            desc2 = tools2_dict[name].strip()
            desc_same = desc1 == desc2
            
            # Check for whitespace-only differences
            desc1_raw = tools1_dict[name]
            desc2_raw = tools2_dict[name]
            whitespace_only = (desc1 == desc2 and desc1_raw != desc2_raw)
            
            if desc_same:
                if whitespace_only:
                    notes = "Whitespace differs (spaces/indentation)"
                    status = "modified"
                else:
                    notes = "Same in both"
                    status = "same"
            else:
                notes = "Description differs"
                status = "modified"
        else:
            desc_same = None
            if not in_file1:
                notes = "Only in File2"
                status = "added"
            else:
                notes = "Only in File1"
                status = "removed"
        
        comparison_data.append({
            "name": name,
            "in_file1": in_file1,
            "in_file2": in_file2,
            "desc_same": desc_same if not whitespace_only else False,
            "notes": notes,
            "status": status
        })
        
        # Differences with word-level diff
        desc1 = tools1_dict.get(name, "")
        desc2 = tools2_dict.get(name, "")
        if desc1.strip() != desc2.strip():
            if not desc1:
                change_type = "Added in File2"
                diff1 = []
                diff2 = [{"text": desc2, "type": "added"}]
            elif not desc2:
                change_type = "Removed from File2"
                diff1 = [{"text": desc1, "type": "removed"}]
                diff2 = []
            else:
                change_type = "Modified"
                diff1, diff2 = get_word_diff(desc1, desc2)
            
            differences_data.append({
                "name": name,
                "file1_desc": desc1[:500] + "..." if len(desc1) > 500 else desc1,
                "file2_desc": desc2[:500] + "..." if len(desc2) > 500 else desc2,
                "file1_diff": diff1,
                "file2_diff": diff2,
                "change_type": change_type
            })
    
    for idx, tool in enumerate(normalized_tools1, 1):
        file1_tools_data.append({
            "index": idx,
            "name": tool["name"],
            "description": tool["description"][:500] + "..." if len(tool["description"]) > 500 else tool["description"]
        })
    
    for idx, tool in enumerate(normalized_tools2, 1):
        file2_tools_data.append({
            "index": idx,
            "name": tool["name"],
            "description": tool["description"][:500] + "..." if len(tool["description"]) > 500 else tool["description"]
        })
    
    return {
        "file1_tools": len(tools1_dict),
        "file2_tools": len(tools2_dict),
        "same_count": same_count,
        "modified_count": modified_count,
        "added_count": added_count,
        "removed_count": removed_count,
        "preview_data": {
            "comparison": comparison_data,
            "differences": differences_data,
            "file1_tools": file1_tools_data,
            "file2_tools": file2_tools_data
        }
    }

# ============== AUTH ENDPOINTS ==============

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    """Exchange session_id for session_token and user data."""
    try:
        body = await request.json()
        session_id = body.get("session_id")
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id required")
        
        # Call Emergent auth service
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
        
        if auth_response.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        user_data = auth_response.json()
        
        # Create or update user
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        existing_user = await db.users.find_one({"email": user_data["email"]}, {"_id": 0})
        
        if existing_user:
            user_id = existing_user["user_id"]
        else:
            await db.users.insert_one({
                "user_id": user_id,
                "email": user_data["email"],
                "name": user_data["name"],
                "picture": user_data.get("picture"),
                "created_at": datetime.now(timezone.utc)
            })
        
        # Create session
        session_token = user_data["session_token"]
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        await db.user_sessions.delete_many({"user_id": user_id})
        await db.user_sessions.insert_one({
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": expires_at,
            "created_at": datetime.now(timezone.utc)
        })
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            path="/",
            secure=True,
            httponly=True,
            samesite="none",
            max_age=7 * 24 * 60 * 60
        )
        
        return {
            "user_id": user_id,
            "email": user_data["email"],
            "name": user_data["name"],
            "picture": user_data.get("picture")
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Session creation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current user data."""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user."""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out"}

# ============== HISTORY ENDPOINTS ==============

@api_router.get("/history")
async def get_history(request: Request):
    """Get comparison history for current user."""
    user = await get_current_user(request)
    
    if user:
        # User is logged in - get from database
        history = await db.comparison_history.find(
            {"user_id": user.user_id},
            {"_id": 0}
        ).sort("timestamp", -1).limit(10).to_list(10)
        return {"history": history, "source": "database"}
    else:
        # Not logged in - return empty (frontend uses localStorage)
        return {"history": [], "source": "localStorage"}

@api_router.post("/history")
async def save_history(request: Request, data: SaveHistoryRequest):
    """Save comparison to history."""
    user = await get_current_user(request)
    
    history_id = f"hist_{uuid.uuid4().hex[:12]}"
    history_item = {
        "id": history_id,
        "user_id": user.user_id if user else None,
        "file1_name": data.file1_name,
        "file2_name": data.file2_name,
        "compare_type": data.compare_type,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "file1_tools": data.file1_tools,
        "file2_tools": data.file2_tools,
        "same_count": data.same_count,
        "modified_count": data.modified_count,
        "added_count": data.added_count,
        "removed_count": data.removed_count,
        "preview_data": data.preview_data
    }
    
    if user:
        await db.comparison_history.insert_one(history_item)
        
        # Keep only last 10
        count = await db.comparison_history.count_documents({"user_id": user.user_id})
        if count > 10:
            oldest = await db.comparison_history.find(
                {"user_id": user.user_id},
                {"_id": 1}
            ).sort("timestamp", 1).limit(count - 10).to_list(count - 10)
            if oldest:
                await db.comparison_history.delete_many(
                    {"_id": {"$in": [doc["_id"] for doc in oldest]}}
                )
    
    return {"id": history_id, "saved": True}

@api_router.delete("/history/{history_id}")
async def delete_history(history_id: str, request: Request):
    """Delete a history item."""
    user = await get_current_user(request)
    if user:
        await db.comparison_history.delete_one({"id": history_id, "user_id": user.user_id})
    return {"deleted": True}

@api_router.delete("/history")
async def clear_history(request: Request):
    """Clear all history."""
    user = await get_current_user(request)
    if user:
        await db.comparison_history.delete_many({"user_id": user.user_id})
    return {"cleared": True}


# ============== SHARE ENDPOINTS ==============

class ShareComparisonRequest(BaseModel):
    """Request model for sharing a comparison."""
    file1_name: str
    file2_name: str
    compare_type: str
    output_filename: str
    summary: Dict[str, Any]
    preview_data: Dict[str, Any]
    download_url: Optional[str] = None

@api_router.post("/share")
async def share_comparison(request: Request, data: ShareComparisonRequest):
    """Save comparison and generate shareable link."""
    share_id = f"share_{uuid.uuid4().hex[:16]}"
    user = await get_current_user(request)
    
    shared_item = {
        "id": share_id,
        "user_id": user.user_id if user else None,
        "file1_name": data.file1_name,
        "file2_name": data.file2_name,
        "compare_type": data.compare_type,
        "output_filename": data.output_filename,
        "summary": data.summary,
        "preview_data": data.preview_data,
        "download_url": data.download_url,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
        "view_count": 0
    }
    
    await db.shared_comparisons.insert_one(shared_item)
    
    # Generate shareable URL
    base_url = os.environ.get('REACT_APP_BACKEND_URL', 'http://localhost:3000')
    share_url = f"{base_url}/shared/{share_id}"
    
    return {
        "share_id": share_id,
        "share_url": share_url,
        "expires_at": shared_item["expires_at"]
    }

@api_router.get("/shared/{share_id}")
async def get_shared_comparison(share_id: str):
    """Get shared comparison data."""
    shared_item = await db.shared_comparisons.find_one(
        {"id": share_id},
        {"_id": 0}
    )
    
    if not shared_item:
        raise HTTPException(status_code=404, detail="Shared comparison not found or expired")
    
    # Check if expired
    expires_at = datetime.fromisoformat(shared_item["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        await db.shared_comparisons.delete_one({"id": share_id})
        raise HTTPException(status_code=410, detail="Shared comparison has expired")
    
    # Increment view count
    await db.shared_comparisons.update_one(
        {"id": share_id},
        {"$inc": {"view_count": 1}}
    )
    
    return shared_item

# ============== API ENDPOINTS ==============

@api_router.get("/")
async def root():
    return {
        "message": "JSON Comparison Tool API",
        "max_file_size": MAX_FILE_SIZE,
        "max_file_size_mb": MAX_FILE_SIZE // (1024 * 1024)
    }

@api_router.get("/config")
async def get_config():
    """Get app configuration."""
    return {
        "max_file_size": MAX_FILE_SIZE,
        "max_file_size_mb": MAX_FILE_SIZE // (1024 * 1024),
        "max_history_items": 10
    }

@api_router.post("/upload", response_model=JsonUploadResponse)
async def upload_json(file: UploadFile = File(...)):
    """Upload and validate a JSON file."""
    file_id = str(uuid.uuid4())
    
    try:
        content = await file.read()
        size = len(content)
        
        # Check file size
        if size > MAX_FILE_SIZE:
            return JsonUploadResponse(
                file_id=file_id,
                filename=file.filename,
                size=size,
                valid=False,
                error=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)} MB"
            )
        
        # Validate JSON
        try:
            data = json.loads(content.decode('utf-8'))
        except json.JSONDecodeError as e:
            return JsonUploadResponse(
                file_id=file_id,
                filename=file.filename,
                size=size,
                valid=False,
                error=f"Invalid JSON: {str(e)} at line {e.lineno}, column {e.colno}"
            )
        
        # Save to temp storage
        file_path = TEMP_DIR / f"{file_id}.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f)
        
        structure = get_json_structure(data)
        
        return JsonUploadResponse(
            file_id=file_id,
            filename=file.filename,
            size=size,
            valid=True,
            structure=structure
        )
    except Exception as e:
        logger.error(f"Error uploading file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/upload-content")
async def upload_json_content(request: Request):
    """Upload JSON content directly (for edit mode)."""
    try:
        body = await request.json()
        content = body.get("content", "")
        filename = body.get("filename", "edited.json")
        
        file_id = str(uuid.uuid4())
        
        # Parse JSON
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            return {
                "file_id": file_id,
                "filename": filename,
                "size": len(content),
                "valid": False,
                "error": f"Invalid JSON: {str(e)} at line {e.lineno}"
            }
        
        # Save to temp storage
        file_path = TEMP_DIR / f"{file_id}.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f)
        
        return {
            "file_id": file_id,
            "filename": filename,
            "size": len(content),
            "valid": True,
            "structure": get_json_structure(data)
        }
    except Exception as e:
        logger.error(f"Error uploading content: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/file-content/{file_id}")
async def get_file_content(file_id: str):
    """Get the JSON content of an uploaded file."""
    file_path = TEMP_DIR / f"{file_id}.json"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return {"content": content}
    except Exception as e:
        logger.error(f"Error reading file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/analyze/{file_id}", response_model=AnalyzeResponse)
async def analyze_json(file_id: str):
    """Analyze JSON structure and detect possible tool paths."""
    file_path = TEMP_DIR / f"{file_id}.json"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        array_paths = find_array_paths(data)
        
        detected_paths = []
        for path in TOOL_PATHS:
            try:
                current = data
                for key in path:
                    current = current[key]
                if isinstance(current, list) and len(current) > 0:
                    detected_paths.append(ToolPath(
                        path=path,
                        path_string=" -> ".join(path),
                        tool_count=len(current)
                    ))
            except (KeyError, TypeError):
                continue
        
        all_paths = {p.path_string: p for p in detected_paths + array_paths}
        structure = get_json_structure(data)
        
        return AnalyzeResponse(
            detected_paths=list(all_paths.values()),
            json_structure=structure
        )
    except Exception as e:
        logger.error(f"Error analyzing file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/tools/{file_id}")
async def get_tools(file_id: str, path: Optional[str] = None):
    """Get list of tools from a JSON file."""
    file_path = TEMP_DIR / f"{file_id}.json"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        tools, found_path = extract_tools(data, path)
        normalized = [normalize_tool(t, i) for i, t in enumerate(tools, 1)]
        
        return {
            "path": found_path,
            "count": len(normalized),
            "tools": [{"name": t["name"], "description": t["description"][:200] + "..." if len(t["description"]) > 200 else t["description"]} for t in normalized]
        }
    except Exception as e:
        logger.error(f"Error getting tools: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/compare", response_model=ComparisonSummary)
async def compare_files(request_data: CompareRequest):
    """Compare two JSON files and generate Excel report."""
    file1_path = TEMP_DIR / f"{request_data.file1_id}.json"
    file2_path = TEMP_DIR / f"{request_data.file2_id}.json"
    
    if not file1_path.exists() or not file2_path.exists():
        raise HTTPException(status_code=404, detail="One or both files not found")
    
    try:
        with open(file1_path, 'r', encoding='utf-8') as f:
            data1 = json.load(f)
        with open(file2_path, 'r', encoding='utf-8') as f:
            data2 = json.load(f)
        
        # Extract tools based on compare type
        if request_data.compare_type == "entire":
            tools1 = [{"name": "Entire JSON", "description": json.dumps(data1, indent=2)}]
            tools2 = [{"name": "Entire JSON", "description": json.dumps(data2, indent=2)}]
        elif request_data.compare_type == "system":
            system_keys = ["system", "systemPrompt", "system_prompt", "config", "configuration"]
            tools1 = []
            tools2 = []
            for key in system_keys:
                if key in data1:
                    val = data1[key]
                    tools1.append({"name": key, "description": json.dumps(val, indent=2) if isinstance(val, (dict, list)) else str(val)})
                if key in data2:
                    val = data2[key]
                    tools2.append({"name": key, "description": json.dumps(val, indent=2) if isinstance(val, (dict, list)) else str(val)})
            if not tools1 and not tools2:
                for key, val in data1.items():
                    tools1.append({"name": key, "description": json.dumps(val, indent=2) if isinstance(val, (dict, list)) else str(val)})
                for key, val in data2.items():
                    tools2.append({"name": key, "description": json.dumps(val, indent=2) if isinstance(val, (dict, list)) else str(val)})
        elif request_data.compare_type == "custom" and request_data.custom_path:
            tools1, _ = extract_tools(data1, request_data.custom_path)
            tools2, _ = extract_tools(data2, request_data.custom_path)
        else:
            tools1, _ = extract_tools(data1)
            tools2, _ = extract_tools(data2)
        
        # Generate Excel
        excel_filename = f"comparison_{uuid.uuid4().hex[:8]}.xlsx"
        excel_path = TEMP_DIR / excel_filename
        
        stats = create_excel_comparison(tools1, tools2, request_data.selected_tools, str(excel_path))
        
        return ComparisonSummary(
            file1_tools=stats["file1_tools"],
            file2_tools=stats["file2_tools"],
            same_count=stats["same_count"],
            modified_count=stats["modified_count"],
            added_count=stats["added_count"],
            removed_count=stats["removed_count"],
            excel_filename=excel_filename,
            download_url=f"/api/download/{excel_filename}",
            preview_data=stats.get("preview_data")
        )
    except Exception as e:
        logger.error(f"Error comparing files: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/download/{filename}")
async def download_excel(filename: str):
    """Download generated Excel file."""
    file_path = TEMP_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/structure/{file_id}")
async def get_full_structure(file_id: str):
    """Get full JSON structure for tree view."""
    file_path = TEMP_DIR / f"{file_id}.json"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return {"structure": data}
    except Exception as e:
        logger.error(f"Error getting structure: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_cleanup():
    pass
